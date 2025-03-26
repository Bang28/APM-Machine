from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from monitoring.models import  Motor, LogActivity
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required(login_url='login')
def download_sample_motor(request):
    # Buat DataFrame dengan format contoh
    sample_data = {
        "Tag No": pd.Series(["72A-M-01"]),
        "Motor Name": pd.Series(["CONVEYOR HYDRA PULPER"]),
        "Output": pd.Series([7.5]),
        "Voltage": pd.Series([380]),
        "Starter Type": pd.Series(["DOL"]),
        "SET OL": pd.Series([15]),
        "Input": pd.Series([15]),
        "Speed": pd.Series([960]),
        "FREK (Boleh Kosong)": pd.Series([16.5, None], dtype="object"),  # Bisa berbeda panjang
        "Frame (Boleh Kosong)": pd.Series(["F100", None], dtype="object"),  # Bisa berbeda panjang
        "Fondation Type": pd.Series(["Rigid"]),
        "Class Type": pd.Series(["Class 2"]),
    }

    # Konversi ke DataFrame tanpa ValueError
    df = pd.DataFrame(dict(sample_data))

    # Simpan ke Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="sample_motor.xlsx"'

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    return response

@login_required(login_url='login')
def import_motor_data(request):
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        try:
            df = pd.read_excel(excel_file, dtype=str)  # Baca semua kolom sebagai string untuk pemrosesan
            
            # Konversi koma ke titik untuk kolom yang seharusnya angka
            numeric_fields = ["Output", "Voltage", "SET OL", "Input", "Speed", "FREK"]
            for field in numeric_fields:
                if field in df.columns:
                    df[field] = df[field].str.replace(",", ".", regex=True)  # Ubah koma ke titik
                    df[field] = pd.to_numeric(df[field], errors="coerce")  # Konversi ke float

            # Validasi: Pastikan hanya "FREK" dan "Frame" yang boleh kosong
            required_fields = ["Tag No", "Motor Name", "Output", "Voltage", "Starter Type", "SET OL", "Input", "Speed", "Fondation Type", "Class Type"]
            missing_values = df[required_fields].isnull().any(axis=1)  # Cek apakah ada kolom wajib yang kosong
            
            if missing_values.any():
                messages.error(request, "Beberapa baris memiliki data yang tidak lengkap. Hanya 'FREK' dan 'Frame' yang boleh kosong.")
                return redirect("data_motor")

            # Simpan ke database
            for _, row in df.iterrows():
                Motor.objects.update_or_create(
                    tag_number=row["Tag No"],
                    defaults={
                        "name": row["Motor Name"],
                        "output": row["Output"],
                        "voltage": row["Voltage"],
                        "starter_type": row["Starter Type"],
                        "set_ol": row["SET OL"],
                        "input": row["Input"],
                        "speed": row["Speed"],
                        "frek": row["FREK (Boleh Kosong)"] if pd.notna(row["FREK (Boleh Kosong)"]) else None,  # Kosongkan jika NaN
                        "frame": row["Frame (Boleh Kosong)"] if pd.notna(row["Frame (Boleh Kosong)"]) else None,  # Kosongkan jika NaN
                        "foundation_type": row["Fondation Type"],
                        "class_type": row["Class Type"],
                    }
                )

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Add",
                description=f"Menambahkan data master motor dari file Excel."
            )

            messages.success(request, "Data berhasil diimport!")
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")

    return redirect("data_motor")

@login_required(login_url='login')
def export_motor_excel(request):
    # Ambil semua tipe starter unik dari database
    starter_types = Motor.objects.values_list('starter_type', flat=True).distinct()
    
    # Buat workbook baru dan hapus sheet default
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    for starter in starter_types:
        # Buat sheet baru berdasarkan starter_type
        sheet_name = starter if starter else "Unknown"
        ws = wb.create_sheet(title=sheet_name[:31])  # Nama sheet maksimal 31 karakter
        
        # Header
        headers = ["NO", "TAG NO", "MOTOR NAME", "OUTPUT (KW)", "VOLTAGE (V)", "STARTER TYPE", "SET OL",
                   "INPUT (A)", "SPEED (RPM)", "FREK (HZ)", "FRAME", "FOUNDATION TYPE", "CLASS TYPE"]
        ws.append(headers)
        
        # Styling header
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        
        # Ambil data motor berdasarkan starter_type
        motors = Motor.objects.filter(starter_type=starter)
        
        # Tambahkan data ke sheet
        for index, motor in enumerate(motors, start=1):
            ws.append([
                index, motor.tag_number, motor.name, motor.output, motor.voltage, motor.starter_type,
                motor.set_ol, motor.input, motor.speed, motor.frek, motor.frame, motor.foundation_type,
                motor.class_type
            ])
    
    # Buat response HTTP dengan file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data_motors.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
def data_motor(request):
    motors = Motor.objects.all().order_by('-created_at')
    context = {
        'motors': motors
    }
    return render(request, 'tables/data_motor.html', context)

@login_required(login_url='login')
def add_data_motor(request):
    if request.method == "POST":
        try:
            # field yang wajib diisi
            required_fields = [
                'tag_number', 'name', 'output', 'voltage', 'starter_type', 'set_ol', 'input', 'speed', 'class_type', 'foundation_type'
            ]
            # validasi untuk field yang wajib diisi
            if not all(request.POST.get(field) for field in required_fields):
                messages.error(request, 'Semua field wajib diisi kecuali Frame dan Frekuensi!')
                return redirect('data_motor')
            
            # Fungsi bantu untuk parsing float
            def get_float(value, default=0.0):
                try:
                    return float(value) if value else default
                except ValueError:
                    return default
            
            # parsing ke tipe data numerik
            output = get_float(request.POST.get('output'))
            voltage = get_float(request.POST.get('voltage'))
            set_ol = get_float(request.POST.get('set_ol'))
            input_power = get_float(request.POST.get('input'))
            speed = get_float(request.POST.get('speed'))
            frek = get_float(request.POST.get('frek'))

            # frame
            frame = request.POST.get('frame') or None # biarkan kosong jika tidak ada      
            
            motor = Motor(
                tag_number = request.POST.get('tag_number'),
                name = request.POST.get('name'),
                output = output,
                voltage = voltage,
                starter_type = request.POST.get('starter_type'),
                set_ol = set_ol,
                input = input_power,
                speed = speed,
                frek = frek,
                frame = frame,
                foundation_type = request.POST.get('foundation_type'),
                class_type = request.POST.get('class_type')
            )
            motor.save()

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Add",
                description=f"Menambahkan data master motor - Kode Motor : [{motor.tag_number}]."
            )
            
            messages.success(request, 'Data berhasil ditambahkan!')
            return redirect('data_motor')

        except ValueError as ve:
            error_msg = f"Harap masukkan angka yang valid untuk field numerik! Error {ve}"
            print(f"[Error ] {error_msg}")
            messages.error(request, error_msg)
        
        except Exception as e:
            error_msg = f"Terjadi error {e}"
            print(f"[Error ] {error_msg}")
            messages.error(request, error_msg)

        return redirect('data_motor')
    
@login_required(login_url='login')
def edit_data_motor(request, id):
    motor = get_object_or_404(Motor, id=id)

    if request.method == "POST":
        try:
            motor.tag_number = request.POST.get('tag_number')
            motor.name = request.POST.get('name')
            motor.output = float(request.POST.get('output')) if request.POST.get('output') else None
            motor.voltage = float(request.POST.get('voltage')) if request.POST.get('voltage') else None
            motor.starter_type = request.POST.get('starter_type')
            motor.set_ol = float(request.POST.get('set_ol')) if request.POST.get('set_ol') else None
            motor.input = float(request.POST.get('input')) if request.POST.get('input') else None
            motor.speed = float(request.POST.get('speed')) if request.POST.get('speed') else None
            motor.frek = float(request.POST.get('frek')) if request.POST.get('frek') else None
            motor.frame = request.POST.get('frame')
            motor.foundation_type = request.POST.get('foundation_type')
            motor.class_type = request.POST.get('class_type')

            motor.save()

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Change",
                description=f"Memperbarui data master motor - Kode Motor : [{motor.tag_number}]."
            )
            
            messages.success(request, "Data berhasil diperbarui!")
            return redirect('data_motor')
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('data_motor')
    
@login_required(login_url='login')    
def del_data_motor(request, id):
    motor = Motor.objects.get(id=id)
    motor.delete()

    # Simpan log aktivitas
    LogActivity.objects.create(
        user=request.user,
        action="Delete",
        description=f"Menghapus data master motor - Kode Motor : [{motor.tag_number}]."
    )
    messages.success(request, "Data berhasil dihapus!")
    return redirect('data_motor')