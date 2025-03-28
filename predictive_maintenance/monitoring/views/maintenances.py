from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from django.contrib.auth.decorators import login_required
from monitoring.models import  Motor, Measurement, Vibration, Maintenance, LogActivity
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from collections import defaultdict

@login_required(login_url='login')
def export_maintenance_excel(request):
    # Ambil parameter filter dari URL
    tag_no = request.GET.get("tag_no", "").strip()  # Ambil kode motor dari URL

    # Filter data berdasarkan kode motor (jika diberikan)
    if tag_no:
        maintenances = Maintenance.objects.filter(motor__tag_number__icontains=tag_no)
    else:
        maintenances = Maintenance.objects.all()

    # Jika tidak ada data, kembalikan response kosong
    if not maintenances.exists():
        return HttpResponse("Tidak ada data yang ditemukan untuk filter yang diberikan.", status=404)

    # **1. Mengelompokkan data berdasarkan bulan**
    monthly_data = defaultdict(list)
    for maintenance in maintenances:
        month_key = maintenance.tanggal_maintenance.strftime('%m-%Y')  # Format "03-2025"
        monthly_data[month_key].append(maintenance)

    # Buat workbook baru
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Hapus sheet default

    # Loop untuk setiap bulan
    for month_key, maintenances_in_month in monthly_data.items():
        sheet_name = month_key  # Misal: "03-2025"
        ws = wb.create_sheet(title=sheet_name)

        # Header kolom
        headers = [
            "NO", "TAG NO", "MOTOR NAME", "OUT PUT (KW)", "VOLT (V)", "TYPE", "SET OL", "Speed (RPM)",
            "FREK (HZ)", "LOAD CURRENT (IB) R (A)", "LOAD CURRENT (IB) S (A)", "LOAD CURRENT (IB) T (A)",
            "BEARING TEMP D.E (Max 80°C)", "BEARING TEMP N.D.E (Max 80°C)", "COIL TEMP (Max 80°C)",
            "VIBRASI MOTOR (FOUNDATION) RIGID D.E (Max 4,5 mm/s)", "VIBRASI MOTOR (FOUNDATION) RIGID N.D.E (Max 4,5 mm/s)",
            "VIBRASI MOTOR (FOUNDATION) FLEXIBLE D.E (Max 7,1 mm/s)", "VIBRASI MOTOR (FOUNDATION) FLEXIBLE N.D.E (Max 7,1 mm/s)",
            "VIBRASI BEARING CLASS 2 D.E (Max 4 gE)", "VIBRASI BEARING CLASS 2 N.D.E (Max 10 gE)",
            "VIBRASI BEARING CLASS 3 D.E (Max 4 gE)", "VIBRASI BEARING CLASS 3 N.D.E (Max 10 gE)",
            "KETERANGAN", "TANGGAL MAINTENANCE"
        ]

        # Tambahkan header ke sheet
        ws.append(headers)

        # Styling header
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Loop untuk setiap data maintenance dalam bulan yang sama
        for index, maintenance in enumerate(maintenances_in_month, start=1):
            motor = maintenance.motor
            measurement = Measurement.objects.filter(maintenance=maintenance).first()
            vibration = Vibration.objects.filter(maintenance=maintenance).first()

            # **2. Format tanggal menjadi DD/MM/YYYY**
            formatted_date = maintenance.tanggal_maintenance.strftime('%d/%m/%Y')

            ws.append([
                index, motor.tag_number, motor.name, motor.output, motor.voltage, motor.starter_type, motor.set_ol,
                motor.speed, motor.frek,
                measurement.load_current_r if measurement else "", measurement.load_current_s if measurement else "",
                measurement.load_current_t if measurement else "",
                measurement.bearing_temp_de if measurement else "", measurement.bearing_temp_nde if measurement else "",
                measurement.coil_temp if measurement else "",
                vibration.vib_rigid_de if vibration else "", vibration.vib_rigid_nde if vibration else "",
                vibration.vib_flexible_de if vibration else "", vibration.vib_flexible_nde if vibration else "",
                vibration.vib_class_2_de if vibration else "", vibration.vib_class_2_nde if vibration else "",
                vibration.vib_class_3_de if vibration else "", vibration.vib_class_3_nde if vibration else "",
                maintenance.keterangan, formatted_date  # **Tambahkan kolom Tanggal Maintenance**
            ])

    # Buat nama file ekspor berdasarkan filter
    filename = "all-data-maintenance-motor.xlsx"
    if tag_no:
        filename = f"data-maintenance-motor-{tag_no.replace(' ', '_')}.xlsx"

    # Simpan workbook ke response HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

@login_required(login_url='login')
def data_maintenance(request):
    motors = Motor.objects.all().order_by('-created_at')
    maintenances = Maintenance.objects.select_related('motor') \
                                      .prefetch_related('measurement', 'vibration') \
                                      .order_by('-created_at')
    context = {
        'motors': motors,
        'maintenances': maintenances
    }
    return render(request, 'tables/data_maintenance.html', context)

@login_required(login_url='login')
def get_motor_details(request, motor_id):
    """Fungsi mengambil detail motor untuk dinamis form di tambah data maintenance"""
    motor = Motor.objects.filter(id=motor_id).first()
    # Normalisasi format string sebelum dikirim ke frontend
    class_type = motor.class_type.capitalize() if motor.class_type else None
    foundation_type = motor.foundation_type.capitalize() if motor.foundation_type else None
    if motor:
        return JsonResponse({'class_type': class_type, 'foundation_type': foundation_type,})
    return JsonResponse({'error': 'Motor not found'}, status=404)

@login_required(login_url='login')
def add_data_maintenance(request):
    if request.method == "POST":
        try:
            motor_id = request.POST.get('motor')

            # Cek apakah motor dipilih
            if not motor_id:
                messages.error(request, "Motor harus dipilih.")
                return redirect('data_maintenance')

            try:
                motor = Motor.objects.get(id=motor_id)
            except Motor.DoesNotExist:
                messages.error(request, "Motor tidak ditemukan.")
                return redirect('data_maintenance')

            # Fungsi bantu untuk parsing float
            def get_float(value, default=0.0):
                try:
                    return float(value) if value else default
                except ValueError:
                    return default

            # Parsing ke tipe data float
            lc_r = get_float(request.POST.get('load_current_r'))
            lc_s = get_float(request.POST.get('load_current_s'))
            lc_t = get_float(request.POST.get('load_current_t'))
            bt_de = get_float(request.POST.get('bearing_temp_de'))
            bt_nde = get_float(request.POST.get('bearing_temp_nde'))
            coil = get_float(request.POST.get('coil_temp'))

            vr_de = get_float(request.POST.get('vib_rigid_de'))
            vr_nde = get_float(request.POST.get('vib_rigid_nde'))
            vf_de = get_float(request.POST.get('vib_flexible_de'))
            vf_nde = get_float(request.POST.get('vib_flexible_nde'))
            vc2_de = get_float(request.POST.get('vib_class_2_de'))
            vc2_nde = get_float(request.POST.get('vib_class_2_nde'))
            vc3_de = get_float(request.POST.get('vib_class_3_de'))
            vc3_nde = get_float(request.POST.get('vib_class_3_nde'))

            periodik = request.POST.get('periodik', '').strip()
            tgl_main = request.POST.get('tanggal_maintenance')
            keterangan = request.POST.get('keterangan', '').strip()
            pelaksana = request.POST.get('pelaksana', '').strip()

            with transaction.atomic():
                # Simpan data maintenance terlebih dahulu
                maintenance = Maintenance.objects.create(
                    motor=motor,
                    periodik=periodik,
                    tanggal_maintenance = tgl_main,
                    keterangan=keterangan,
                    pelaksana=pelaksana
                )

                # Simpan data measurement yang terkait dengan maintenance
                measurement = Measurement.objects.create(
                    maintenance=maintenance,
                    motor=motor,
                    load_current_r=lc_r,
                    load_current_s=lc_s,
                    load_current_t=lc_t,
                    bearing_temp_de=bt_de,
                    bearing_temp_nde=bt_nde,
                    coil_temp=coil
                )

                # Simpan data vibration yang terkait dengan maintenance
                vibration = Vibration.objects.create(
                    maintenance=maintenance,
                    motor=motor,
                    vib_rigid_de=vr_de,
                    vib_rigid_nde=vr_nde,
                    vib_flexible_de=vf_de,
                    vib_flexible_nde=vf_nde,
                    vib_class_2_de=vc2_de,
                    vib_class_2_nde=vc2_nde,
                    vib_class_3_de=vc3_de,
                    vib_class_3_nde=vc3_nde,
                )

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Add",
                description=f"Menambahkan data maintenance motor - Kode Motor : [{motor.tag_number}]."
            )

            messages.success(request, 'Data maintenance berhasil ditambahkan!')
            return redirect('data_maintenance')

        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('data_maintenance')

@login_required(login_url='login')
def edit_data_maintenance(request, maintenance_id):
    # Ambil data maintenance berdasarkan ID
    maintenance = get_object_or_404(Maintenance, id=maintenance_id)

    # Mengambil Measurement & Vibration terbaru untuk motor ini
    measurement = Measurement.objects.filter(motor=maintenance.motor).order_by('-id').first()
    vibration = Vibration.objects.filter(motor=maintenance.motor).order_by('-id').first()

    # Jika tidak ada data Measurement/Vibration, beri pesan error
    if not measurement:
        messages.error(request, "Data Measurement tidak ditemukan.")
        return redirect('data_maintenance')

    if not vibration:
        messages.error(request, "Data Vibration tidak ditemukan.")
        return redirect('data_maintenance')

    if request.method == "POST":
        try:
            # Fungsi bantu untuk parsing float
            def get_float(value, default=0.0):
                try:
                    return float(value) if value else default
                except ValueError:
                    return default

            # Parsing nilai baru dari request
            lc_r = get_float(request.POST.get('load_current_r'))
            lc_s = get_float(request.POST.get('load_current_s'))
            lc_t = get_float(request.POST.get('load_current_t'))
            bt_de = get_float(request.POST.get('bearing_temp_de'))
            bt_nde = get_float(request.POST.get('bearing_temp_nde'))
            coil = get_float(request.POST.get('coil_temp'))

            vr_de = get_float(request.POST.get('vib_rigid_de'))
            vr_nde = get_float(request.POST.get('vib_rigid_nde'))
            vf_de = get_float(request.POST.get('vib_flexible_de'))
            vf_nde = get_float(request.POST.get('vib_flexible_nde'))
            vc2_de = get_float(request.POST.get('vib_class_2_de'))
            vc2_nde = get_float(request.POST.get('vib_class_2_nde'))
            vc3_de = get_float(request.POST.get('vib_class_3_de'))
            vc3_nde = get_float(request.POST.get('vib_class_3_nde'))

            ket = request.POST.get('keterangan')

            with transaction.atomic():
                # Update data Measurement
                measurement.load_current_r = lc_r
                measurement.load_current_s = lc_s
                measurement.load_current_t = lc_t
                measurement.bearing_temp_de = bt_de
                measurement.bearing_temp_nde = bt_nde
                measurement.coil_temp = coil
                measurement.full_clean()
                measurement.save()

                # Update data Vibration
                vibration.vib_rigid_de = vr_de 
                vibration.vib_rigid_nde = vr_nde 
                vibration.vib_flexible_de = vf_de 
                vibration.vib_flexible_nde = vf_nde 
                vibration.vib_class_2_de = vc2_de
                vibration.vib_class_2_nde = vc2_nde
                vibration.vib_class_3_de = vc3_de
                vibration.vib_class_3_nde = vc3_nde
                vibration.full_clean()
                vibration.save()

                # Update data maintenance
                maintenance.keterangan = ket
                maintenance.full_clean()
                maintenance.save()

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Change",
                description=f"Memperbarui data maintenance motor - Kode Motor : [{maintenance.motor.tag_number}]."
            )

            messages.success(request, 'Data maintenance berhasil diperbarui!')
            return redirect('data_maintenance')

        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('data_maintenance')

@login_required(login_url='login')
def del_data_maintenance(request, maintenance_id):
    if request.method == "POST": # Hanya izinkan POST untuk penghapusan
        try:
            with transaction.atomic():
                maintenance = get_object_or_404(Maintenance, id=maintenance_id)

                # Hapus hanya measurement dan vibration yang terkait dengan maintenance ini
                Measurement.objects.filter(maintenance=maintenance).delete()
                Vibration.objects.filter(maintenance=maintenance).delete()

                # Hapus maintenance
                maintenance.delete()

            # Simpan log aktivitas
            LogActivity.objects.create(
                user=request.user,
                action="Delete",
                description=f"Menghapus data maintenance motor - Kode Motor : [{maintenance.motor.tag_number}]."
            )

            messages.success(request, 'Data maintenance berhasil dihapus!')
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
        
        return redirect('data_maintenance')

    return redirect('data_maintenance')